import logging
from openpyxl.styles import PatternFill, Font, Alignment
from config import STATUS, CURRENT_DATETIME, CURRENT_USER

logger = logging.getLogger(__name__)

def format_comparison_sheet(wb, comparison_df):
    """Format the comparison analysis sheet.
    
    Args:
        wb: Excel workbook object
        comparison_df (DataFrame): Comparison analysis data
    """
    sheet = wb['Comparison Analysis']
    
    # Define column widths
    sheet.column_dimensions['A'].width = 15  # Item Code
    sheet.column_dimensions['B'].width = 30  # Description
    
    # Find column indices
    status_col_idx = None
    diff_col_idx = None
    net_diff_col_idx = None
    annual_forecast_col_idx = None
    adjusted_forecast_col_idx = None
    available_stock_col_idx = None
    pending_orders_col_idx = None
    
    for idx, cell in enumerate(sheet[1]):
        if cell.value == 'Status':
            status_col_idx = idx + 1
        elif cell.value == 'Difference':
            diff_col_idx = idx + 1
        elif cell.value == 'Net Difference':
            net_diff_col_idx = idx + 1
        elif cell.value == 'Annual Forecast':
            annual_forecast_col_idx = idx + 1
        elif cell.value == 'Adjusted Annual Forecast':
            adjusted_forecast_col_idx = idx + 1
        elif cell.value == 'Available Stock':
            available_stock_col_idx = idx + 1
        elif cell.value == 'Pending Orders':
            pending_orders_col_idx = idx + 1
    
    # Format headers
    for idx in [status_col_idx, diff_col_idx, net_diff_col_idx, annual_forecast_col_idx, 
                adjusted_forecast_col_idx, available_stock_col_idx, pending_orders_col_idx]:
        if idx:
            sheet.cell(row=1, column=idx).font = Font(bold=True)
    
    # Format status cells
    if status_col_idx:
        for row in range(2, len(comparison_df) + 2):
            cell = sheet.cell(row=row, column=status_col_idx)
            
            if cell.value == "HIGH_DEVIATION":
                cell.fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                      end_color=STATUS['HIGH_DEVIATION']['color'], 
                                      fill_type="solid")
            elif cell.value == "MODERATE_DEVIATION":
                cell.fill = PatternFill(start_color=STATUS['MODERATE_DEVIATION']['color'], 
                                      end_color=STATUS['MODERATE_DEVIATION']['color'], 
                                      fill_type="solid")
            elif cell.value == "ACCEPTABLE":
                cell.fill = PatternFill(start_color=STATUS['ACCEPTABLE']['color'], 
                                      end_color=STATUS['ACCEPTABLE']['color'], 
                                      fill_type="solid")
            elif cell.value == "LOW_REQUEST":
                cell.fill = PatternFill(start_color=STATUS['LOW_REQUEST']['color'], 
                                      end_color=STATUS['LOW_REQUEST']['color'], 
                                      fill_type="solid")
            elif cell.value == "COVERED_BY_STOCK":
                cell.fill = PatternFill(start_color=STATUS['COVERED_BY_STOCK']['color'], 
                                      end_color=STATUS['COVERED_BY_STOCK']['color'], 
                                      fill_type="solid")
            elif cell.value == "COVERED_BY_ORDERS":
                cell.fill = PatternFill(start_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                      end_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                      fill_type="solid")
    
    # Format difference cells
    if diff_col_idx:
        for row in range(2, len(comparison_df) + 2):
            cell = sheet.cell(row=row, column=diff_col_idx)
            value = cell.value
            
            if value is not None:
                if value > 3:
                    cell.fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                          end_color=STATUS['HIGH_DEVIATION']['color'], 
                                          fill_type="solid")
                elif 1 <= value <= 3:
                    cell.fill = PatternFill(start_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          end_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          fill_type="solid")
                elif value == 0:
                    cell.fill = PatternFill(start_color=STATUS['ACCEPTABLE']['color'], 
                                          end_color=STATUS['ACCEPTABLE']['color'], 
                                          fill_type="solid")
                else:  # value < 0
                    cell.fill = PatternFill(start_color=STATUS['LOW_REQUEST']['color'], 
                                          end_color=STATUS['LOW_REQUEST']['color'], 
                                          fill_type="solid")
    
    # Format net difference cells
    if net_diff_col_idx:
        for row in range(2, len(comparison_df) + 2):
            cell = sheet.cell(row=row, column=net_diff_col_idx)
            value = cell.value
            
            if value is not None:
                if value > 3:
                    cell.fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                          end_color=STATUS['HIGH_DEVIATION']['color'], 
                                          fill_type="solid")
                elif 1 <= value <= 3:
                    cell.fill = PatternFill(start_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          end_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          fill_type="solid")
                elif value == 0:
                    cell.fill = PatternFill(start_color=STATUS['ACCEPTABLE']['color'], 
                                          end_color=STATUS['ACCEPTABLE']['color'], 
                                          fill_type="solid")
                else:  # value < 0
                    cell.fill = PatternFill(start_color=STATUS['LOW_REQUEST']['color'], 
                                          end_color=STATUS['LOW_REQUEST']['color'], 
                                          fill_type="solid")
    
    # Highlight stock and pending orders if they can help fulfill requirements
    if diff_col_idx and available_stock_col_idx and pending_orders_col_idx:
        for row in range(2, len(comparison_df) + 2):
            diff_cell = sheet.cell(row=row, column=diff_col_idx)
            stock_cell = sheet.cell(row=row, column=available_stock_col_idx)
            orders_cell = sheet.cell(row=row, column=pending_orders_col_idx)
            
            diff_value = diff_cell.value if diff_cell.value is not None else 0
            stock_value = stock_cell.value if stock_cell.value is not None else 0
            orders_value = orders_cell.value if orders_cell.value is not None else 0
            
            # If stock alone can cover the difference
            if diff_value > 0 and stock_value >= diff_value:
                stock_cell.fill = PatternFill(start_color=STATUS['COVERED_BY_STOCK']['color'], 
                                            end_color=STATUS['COVERED_BY_STOCK']['color'], 
                                            fill_type="solid")
                stock_cell.font = Font(bold=True)
            
            # If stock + orders can cover the difference
            elif diff_value > 0 and (stock_value + orders_value) >= diff_value:
                orders_cell.fill = PatternFill(start_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                             end_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                             fill_type="solid")
                orders_cell.font = Font(bold=True)
    
    # Highlight adjusted forecasts
    if annual_forecast_col_idx and adjusted_forecast_col_idx:
        for row in range(2, len(comparison_df) + 2):
            annual = sheet.cell(row=row, column=annual_forecast_col_idx).value
            adjusted = sheet.cell(row=row, column=adjusted_forecast_col_idx).value
            
            # If the forecast was adjusted (i.e., value was < 1)
            if annual is not None and adjusted is not None and annual != adjusted:
                sheet.cell(row=row, column=adjusted_forecast_col_idx).font = Font(bold=True, color="0000FF")
    
    # Add a header with explanation
    sheet.insert_rows(1, 2)
    sheet.cell(row=1, column=1).value = "COMPARISON ANALYSIS - Including Stock & Order Consideration"
    sheet.cell(row=1, column=1).font = Font(size=14, bold=True)
    sheet.cell(row=2, column=1).value = f"Generated: {CURRENT_DATETIME} | User: {CURRENT_USER} | Note: Analysis considers both available stock and pending orders"
    sheet.cell(row=2, column=1).font = Font(italic=True)
    
    # Merge cells for the header
    merge_end = min(10, sheet.max_column)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end)