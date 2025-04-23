import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from config import STATUS, CURRENT_DATETIME, CURRENT_USER

def create_dashboard(wb, comparison_df, summary_df, plant_df, unique_plants):
    """Create and format the dashboard sheet.
    
    Args:
        wb: Excel workbook object
        comparison_df (DataFrame): Comparison analysis data
        summary_df (DataFrame): Summary statistics
        plant_df (DataFrame): Plant summary data
        unique_plants (list): List of unique plant names
    """
    
    # Create dashboard sheet if it doesn't exist
    if 'Dashboard' in wb.sheetnames:
        dashboard = wb['Dashboard']
    else:
        dashboard = wb.create_sheet('Dashboard')
    
    # Clear any existing content
    for row in dashboard.iter_rows():
        for cell in row:
            cell.value = None
    
    # Title
    dashboard['A1'] = "GASKET INVENTORY ANALYSIS DASHBOARD"
    dashboard['A1'].font = Font(size=16, bold=True)
    dashboard['A2'] = f"Generated on: {CURRENT_DATETIME} | User: {CURRENT_USER}"
    dashboard['A2'].font = Font(italic=True)
    
    # Notice about forecast adjustment and inventory consideration
    dashboard['A3'] = "Note: Analysis considers available stock and pending orders when determining status"
    dashboard['A3'].font = Font(italic=True, color="0000FF")
    dashboard['A4'] = "Note: Annual Forecast values < 1 are treated as 0 for analysis"
    dashboard['A4'].font = Font(italic=True, color="0000FF")
    
    # ----- HIGH DEVIATION SECTION -----
    dashboard['A5'] = "⚠️ HIGH DEVIATION ITEMS (Net Difference > 3 after stock & orders)"
    dashboard['A5'].font = Font(bold=True, size=12, color="FF0000")
    
    # Get high deviation items that cannot be covered by stock or orders
    high_deviation_items = comparison_df[comparison_df['Status'] == 'HIGH_DEVIATION'].sort_values('Net Difference', ascending=False)
    
    # Headers
    headers = ['Item Code', 'Description', 'Annual Forecast', 'Total Plant Requests', 
               'Available Stock', 'Pending Orders', 'Net Difference', 'Status', 'Recommendation']
    for col, header in enumerate(headers):
        dashboard.cell(row=6, column=col+1).value = header
        dashboard.cell(row=6, column=col+1).font = Font(bold=True)
    
    # Data rows
    if not high_deviation_items.empty:
        display_rows = min(10, len(high_deviation_items))  # Show up to 10 items
        for i in range(display_rows):
            row_data = high_deviation_items.iloc[i]
            dashboard.cell(row=7+i, column=1).value = row_data['Item Code']
            dashboard.cell(row=7+i, column=2).value = row_data['Description']
            dashboard.cell(row=7+i, column=3).value = row_data['Annual Forecast']
            dashboard.cell(row=7+i, column=4).value = row_data['Total Plant Requests']
            dashboard.cell(row=7+i, column=5).value = row_data['Available Stock']
            dashboard.cell(row=7+i, column=6).value = row_data['Pending Orders']
            dashboard.cell(row=7+i, column=7).value = row_data['Net Difference']
            dashboard.cell(row=7+i, column=8).value = row_data['Status']
            dashboard.cell(row=7+i, column=9).value = row_data['Recommendation']
            
            # Highlight the difference
            dashboard.cell(row=7+i, column=7).fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                                             end_color=STATUS['HIGH_DEVIATION']['color'], 
                                                             fill_type="solid")
        
        if len(high_deviation_items) > 10:
            dashboard.cell(row=17, column=1).value = f"... and {len(high_deviation_items) - 10} more high deviation items"
    else:
        dashboard.cell(row=7, column=1).value = "No items with high deviation found"
    
    # ----- COVERED BY STOCK/ORDERS SECTION -----
    covered_row = 19
    dashboard.cell(row=covered_row, column=1).value = "✅ ITEMS COVERED BY STOCK OR PENDING ORDERS"
    dashboard.cell(row=covered_row, column=1).font = Font(bold=True, size=12, color="009900")
    
    # Get items covered by stock or orders
    covered_items = pd.concat([
        comparison_df[comparison_df['Status'] == 'COVERED_BY_STOCK'],
        comparison_df[comparison_df['Status'] == 'COVERED_BY_ORDERS']
    ]).sort_values('Difference', ascending=False)
    
    # Headers
    headers = ['Item Code', 'Description', 'Annual Forecast', 'Total Plant Requests', 
               'Available Stock', 'Pending Orders', 'Difference', 'Status', 'Recommendation']
    for col, header in enumerate(headers):
        dashboard.cell(row=covered_row+1, column=col+1).value = header
        dashboard.cell(row=covered_row+1, column=col+1).font = Font(bold=True)
    
    # Create rest of the dashboard with status breakdowns, summaries, etc.
    # (Implementation abbreviated for brevity - see full function in original code)