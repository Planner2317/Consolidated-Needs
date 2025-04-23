import logging
from openpyxl.styles import PatternFill, Font, Alignment
from config import STATUS, CURRENT_DATETIME, CURRENT_USER

logger = logging.getLogger(__name__)

def format_plant_sheet(wb, plant, sheet_name=None):
    """Format the plant sheet with color coding and highlighting.
    
    Args:
        wb: Excel workbook object
        plant (str): Plant name
        sheet_name (str, optional): Sheet name (if different from default)
    """
    # Use provided sheet_name or generate one
    if not sheet_name:
        short_plant_name = plant[:25] if len(plant) > 25 else plant
        sheet_name = f'Plant_{short_plant_name}'
    
    if sheet_name not in wb.sheetnames:
        logger.warning(f"Warning: Sheet '{sheet_name}' not found")
        return
    
    sheet = wb[sheet_name]
    
    # Define column widths
    sheet.column_dimensions['A'].width = 15  # Item Code
    
    # Find columns
    plant_status_col_idx = None
    plant_diff_col_idx = None
    plant_net_diff_col_idx = None
    available_stock_col_idx = None
    pending_orders_col_idx = None
    
    for idx, cell in enumerate(sheet[1]):
        if cell.value == 'Plant Status':
            plant_status_col_idx = idx + 1
        elif cell.value == 'Plant Difference':
            plant_diff_col_idx = idx + 1
        elif cell.value == 'Plant Net Difference':
            plant_net_diff_col_idx = idx + 1
        elif cell.value == 'Available Stock':
            available_stock_col_idx = idx + 1
        elif cell.value == 'Pending Orders':
            pending_orders_col_idx = idx + 1
    
    # Format headers
    for i, cell in enumerate(sheet[1]):
        sheet.cell(row=1, column=i+1).font = Font(bold=True)
    
    # Format status cells
    if plant_status_col_idx:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=plant_status_col_idx)
            
            if cell.value == "HIGH_DEVIATION":
                cell.fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                      end_color=STATUS['HIGH_DEVIATION']['color'], 
                                      fill_type="solid")
            elif cell.value == "MODERATE_DEVIATION":
                cell.fill = PatternFill(start_color=STATUS['MODERATE_DEVIATION']['color'], 
                                      end_color=STATUS['MODERATE_DEVIATION']['color'], 
                                      fill_type="solid")
            elif cell.value == "ACCEPTABLE":
                cell.fill = PatternFill(start_color=STATUS['ACCEPTABLE']['color'], 
                                      end_color=STATUS['ACCEPTABLE']['color'], 
                                      fill_type="solid")
            elif cell.value == "LOW_REQUEST":
                cell.fill = PatternFill(start_color=STATUS['LOW_REQUEST']['color'], 
                                      end_color=STATUS['LOW_REQUEST']['color'], 
                                      fill_type="solid")
            elif cell.value == "COVERED_BY_STOCK":
                cell.fill = PatternFill(start_color=STATUS['COVERED_BY_STOCK']['color'], 
                                      end_color=STATUS['COVERED_BY_STOCK']['color'], 
                                      fill_type="solid")
            elif cell.value == "COVERED_BY_ORDERS":
                cell.fill = PatternFill(start_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                      end_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                      fill_type="solid")
    
    # Format difference cells
    if plant_diff_col_idx:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=plant_diff_col_idx)
            value = cell.value
            
            if value is not None:
                if value > 3:
                    cell.fill = PatternFill(start_color=STATUS['HIGH_DEVIATION']['color'], 
                                          end_color=STATUS['HIGH_DEVIATION']['color'], 
                                          fill_type="solid")
                elif 1 <= value <= 3:
                    cell.fill = PatternFill(start_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          end_color=STATUS['MODERATE_DEVIATION']['color'], 
                                          fill_type="solid")
                elif value == 0:
                    cell.fill = PatternFill(start_color=STATUS['ACCEPTABLE']['color'], 
                                          end_color=STATUS['ACCEPTABLE']['color'], 
                                          fill_type="solid")
                else:  # value < 0
                    cell.fill = PatternFill(start_color=STATUS['LOW_REQUEST']['color'], 
                                          end_color=STATUS['LOW_REQUEST']['color'], 
                                          fill_type="solid")
    
    # Highlight stock and pending orders if they can help fulfill requirements
    if plant_diff_col_idx and available_stock_col_idx and pending_orders_col_idx:
        for row in range(2, sheet.max_row + 1):
            diff_cell = sheet.cell(row=row, column=plant_diff_col_idx)
            stock_cell = sheet.cell(row=row, column=available_stock_col_idx)
            orders_cell = sheet.cell(row=row, column=pending_orders_col_idx)
            
            diff_value = diff_cell.value if diff_cell.value is not None else 0
            stock_value = stock_cell.value if stock_cell.value is not None else 0
            orders_value = orders_cell.value if orders_cell.value is not None else 0
            
            # If stock alone can cover the difference
            if diff_value > 0 and stock_value >= diff_value:
                stock_cell.fill = PatternFill(start_color=STATUS['COVERED_BY_STOCK']['color'], 
                                            end_color=STATUS['COVERED_BY_STOCK']['color'], 
                                            fill_type="solid")
                stock_cell.font = Font(bold=True)
            
            # If stock + orders can cover the difference
            elif diff_value > 0 and (stock_value + orders_value) >= diff_value:
                orders_cell.fill = PatternFill(start_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                             end_color=STATUS['COVERED_BY_ORDERS']['color'], 
                                             fill_type="solid")
                orders_cell.font = Font(bold=True)
    
    # Add a title at the top
    sheet.insert_rows(1, 3)
    sheet.cell(row=1, column=1).value = f"PLANT COMMUNICATION SHEET - {plant}"
    sheet.cell(row=1, column=1).font = Font(size=14, bold=True)
    sheet.cell(row=2, column=1).value = f"Generated: {CURRENT_DATETIME} | User: {CURRENT_USER}"
    sheet.cell(row=2, column=1).font = Font(italic=True)
    
    # Add explanation for inventory consideration
    sheet.cell(row=3, column=1).value = "Note: Analysis considers available stock and pending orders when determining status and recommendations"
    sheet.cell(row=3, column=1).font = Font(italic=True, color="0000FF")
    
    # Merge cells for the title
    for row in [1, 2, 3]:
        merge_end = min(10, sheet.max_column)  # Merge up to 10 columns or max columns
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left')